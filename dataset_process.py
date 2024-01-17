import numpy as np
import random
import torch
from torch_geometric.data import InMemoryDataset, Data
from openpyxl import load_workbook

    
class MyOwnDataset(InMemoryDataset):
    def __init__(self, root, test = True, transform=None, pre_transform=None):
        if test:
            self.order = 1
        else:
            self.order = 0
        super(MyOwnDataset, self).__init__(root, transform, pre_transform)
        self.data, self.slices = torch.load(self.processed_paths[0]) 


    @property
    def raw_file_names(self):
        return ['train_validation', 'test']
    @property
    def processed_file_names(self):
        return ['data.pt']
    
    def read_data(self, file):
        wb = load_workbook('./data/'+ file + '.xlsx')
        ws = wb['Sheet1']
        data_list = []
        for i in range(2, ws.max_row + 1):
            one_hot_vec = [0, 0, 0, 0] #one hot encode
            if ws.cell(row= i, column= 18).value != None:
                if i != 2:
                    X = np.stack(x_list, axis= 0)
            
                    x_data = torch.from_numpy(X.copy())
                    edge_list_1 = []
                    edge_list_2 = []
            
                    if n > 1:
                        for ii in range(n):
                            for jj in range(ii):
                                edge_list_1.append(ii)
                                edge_list_1.append(jj)
                                edge_list_2.append(jj)
                                edge_list_2.append(ii)
                                pos_list.append(np.outer(X[jj, :], X[ii, :]))
                                   
                    else:
                        pos_list.append(np.zeros((17, 17)))
            
                    pos = np.stack(pos_list, axis= 0)
                    pos_data = torch.from_numpy(pos.copy())
                    Edge_index = torch.from_numpy(np.stack((np.array(edge_list_1), np.array(edge_list_2))))
                    data_list.append(Data(x= x_data, edge_index= Edge_index, y = Y, pos = pos_data))

                n = 1        
                x_list = []
                pos_list = []
        
                one_hot_vec[ws.cell(row= i, column=3).value - 1] = 1
                x_list.append(one_hot_vec + [ws.cell(row= i, column=j).value for j in range(4, 17)])
                Y = torch.tensor([[ws.cell(row= i, column=j).value for j in range(18, 25)]])
           
            else:
                n += 1
                try:
                    one_hot_vec[ws.cell(row= i, column=3).value - 1] = 1
                    x_list.append(one_hot_vec + [float(ws.cell(row= i, column=m).value) for m in range(4, 17)])
                except:
                    pass

        random.shuffle(data_list)   
        return data_list     

    def process(self):
        
        file = self.raw_file_names[self.order]
        data_list = self.read_data(file)
        data, slices = self.collate(data_list)
        torch.save((data, slices), self.processed_paths[0])

        if self.pre_filter is not None:
            data_list = [data for data in data_list if self.pre_filter(data)]

        if self.pre_transform is not None:
            data_list = [self.pre_transform(data) for data in data_list]



data = MyOwnDataset('train_val_dataset', test=False)

data_test = MyOwnDataset('test_dataset', test=True)

if __name__ == '__main__':
    print(data)

    print("Dataset type: ", type(data))
    print("Dataset features:", data.num_features)
    print("Dataset target:", data.num_classes)
    print("Dataset sample:", data[1])
    print("Dataset Size:", len(data))
    print("Dataset Size:", len(data_test))
